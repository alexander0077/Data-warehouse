import shutil
from math import floor
import os
import pandas
import random
import openpyxl

from excel.Drivers import Drivers
from utils.ClassesUtils import generateBuses as getBuses
from utils.ClassesUtils import generateCoursesAndRides as getCoursesAndRides
from utils.ClassesUtils import getRandomizedTime
from excel.Controllers import Controllers

pesels = []
mockData = pandas.read_csv("MOCK_DATA.csv")
controllersExcel = []
driversExcel = []


def generate_pesel():
    def generate_random_pesel():
        year = random.randint(0, 99)
        month = random.randint(1, 12)
        day = random.randint(1, 28)  # Ignoring February 29 for simplicity
        random_suffix = str(random.randint(1000, 9999))
        return f"{year:02d}{month:02d}{day:02d}{random_suffix}"

    new_pesel = generate_random_pesel()
    while new_pesel in pesels:
        new_pesel = generate_random_pesel()
    pesels.append(new_pesel)
    return new_pesel


def randomSex():
    if random.randint(0, 1000) % 2 == 0:
        return "Male"
    else:
        return "Female"


def ifControlerWasEmployed(startYear, month, controler):
    currentYear = startYear + floor(month / 12)
    currentMonth = (month % 12) + 1
    startYear = int(controler.startEmploymentDate[0:4])
    startMonth = int(controler.startEmploymentDate[5:7])

    if startYear > currentYear:
        return False
    if startYear == currentYear and startMonth > currentMonth:
        return False

    if controler.endEmploymentDate is None:
        return True

    endYear = int(controler.endEmploymentDate[0:4])
    endMonth = int(controler.endEmploymentDate[5:7])

    if endYear > currentYear:
        return True
    if endYear == currentYear and endMonth > currentMonth:
        return True
    return False


def init_Everything():
    # Given worksheet initialize pesels and tables
    # WARNING: if we properly create dbs and sheets, then initializing shouldnt check for duplicate ids and pesels!!!
    for row in ws1.rows:
        contrID = row[0].value
        contrSurname = row[1].value
        contrFirstname = row[2].value
        contrPESEL = row[3].value
        contrZodiac = row[4].value
        contrSex = row[5].value
        contrStartDate = row[6].value
        if len(row) == 8:
            contrEndDate = row[7].value
        else:
            contrEndDate = None
        controllersExcel.append(Controllers(contrID, contrSurname, contrFirstname, contrPESEL, contrZodiac, contrSex,
                                            contrStartDate, contrEndDate))
        pesels.append(contrPESEL)
    for row in ws2.rows:
        driverSurname = row[0].value
        driverFirstname = row[1].value
        driverPESEL = row[2].value
        driverZodiac = row[3].value
        driverSex = row[4].value
        driverStartDate = row[5].value
        driverEndDate = row[6].value
        driversExcel.append(Drivers(driverSurname, driverFirstname, driverPESEL, driverZodiac, driverSex,
                                    driverStartDate, driverEndDate))
        pesels.append(driverPESEL)


def generateRandomDate(start, end):
    return str(random.randint(start, end)) + str(mockData.iloc[random.randint(0, 999)]['date'][4:])


def fireDrivers(percentOfFires, earliestYearToFire, lastYearOfFire):
    numberOfRows = ws2.max_row
    for _ in range(floor(numberOfRows * percentOfFires)):
        fired = True
        while fired is True:
            rowNumber = random.randint(1, numberOfRows)
            if ws2.cell(row=rowNumber, column=7).value is None:
                fired = False

        endEmploymentDate = generateRandomDate(earliestYearToFire, lastYearOfFire)

        ws2.cell(row=rowNumber, column=7, value=endEmploymentDate)

    workBook.save("WorkBook.xlsx")

def fireControlers(percentOfFires, earliestYearToFire, lastYearOfFire):
    numberOfRows = ws1.max_row
    for _ in range(floor(numberOfRows * percentOfFires)):
        fired = True
        while fired is True:
            rowNumber = random.randint(1, numberOfRows)
            if ws1.cell(row=rowNumber, column=8).value is None:
                fired = False

        endEmploymentDate = generateRandomDate(earliestYearToFire, lastYearOfFire)

        ws1.cell(row=rowNumber, column=8, value=endEmploymentDate)

    workBook.save("WorkBook.xlsx")

def generateControllers(yearOfStart, yearOfEnd, numberofRows, alt=False):
    if ws2.max_row == 1:
        contrId = 0
    else:
        contrId = ws2.max_row

    if alt is False:
        tempContr = Controllers(-1, "unknown", "controller", 0, "", "", "", 0)
        ws1.append(tempContr.get_Controller())

    for i in range(numberofRows):
        PESEL = generate_pesel()
        surname = mockData.iloc[random.randint(0, 999)]['surname']
        firstname = mockData.iloc[random.randint(0, 999)]['firstname']
        zodiac_sign = mockData.iloc[random.randint(0, 999)]['zodiac_sign']
        salary = random.randint(10, 40)
        sex = randomSex()
        if alt is False:
            if random.random() < 0.15:  # probability of working from the beginning of the company existence
                startEmploymentDate = str(yearOfStart) + "-01-01"
            else:
                startEmploymentDate = generateRandomDate(yearOfStart, yearOfEnd)
        else:
            startEmploymentDate = generateRandomDate(yearOfStart, yearOfEnd)

        endEmploymentDate = generateRandomDate(int(startEmploymentDate[0:4]), yearOfEnd + 1)
        while startEmploymentDate >= endEmploymentDate:
            endEmploymentDate = generateRandomDate(int(startEmploymentDate[0:4]), yearOfEnd + 1)
        if random.randint(0, 100) % 20 == 0:
            tempContr = Controllers(contrId, surname, firstname, PESEL, zodiac_sign, sex, startEmploymentDate, salary,
                                    endEmploymentDate)
        else:
            tempContr = Controllers(contrId, surname, firstname, PESEL, zodiac_sign, sex, startEmploymentDate, salary)
        controllersExcel.append(tempContr)
        ws1.append(tempContr.get_Controller())
        contrId += 1
    workBook.save("WorkBook.xlsx")


def generateDrivers(yearOfStart, yearOfEnd, numberOfRows, alt=False):
    if alt is False:
        tempContr = Drivers("unknown", "driver", 0, "", "", "")
        ws2.append(tempContr.getDriver())

    for i in range(numberOfRows):
        PESEL = generate_pesel()
        surname = mockData.iloc[random.randint(0, 999)]['surname']
        firstname = mockData.iloc[random.randint(0, 999)]['firstname']
        zodiac_sign = mockData.iloc[random.randint(0, 999)]['zodiac_sign']
        sex = randomSex()
        if alt is False:
            if random.random() < 0.15:  # probability of working from the beginning of the company existence
                startEmploymentDate = str(yearOfStart) + "-01-01"
            else:
                startEmploymentDate = generateRandomDate(yearOfStart, yearOfEnd)
        else:
            startEmploymentDate = generateRandomDate(yearOfStart, yearOfEnd)

        endEmploymentDate = generateRandomDate(int(startEmploymentDate[0:4]), yearOfEnd + 1)
        while startEmploymentDate >= endEmploymentDate:
            endEmploymentDate = generateRandomDate(int(startEmploymentDate[0:4]), yearOfEnd + 1)

        if random.randint(0, 100) % 20 == 0:
            tempContr = Drivers(surname, firstname, PESEL, zodiac_sign, sex, startEmploymentDate,
                                endEmploymentDate)
        else:
            tempContr = Drivers(surname, firstname, PESEL, zodiac_sign, sex, startEmploymentDate)
        driversExcel.append(tempContr)
        ws2.append(tempContr.getDriver())
    workBook.save("WorkBook.xlsx")

def mainLoop():
    while True:
        mainInput = input()
        if mainInput == "help":
            # List all commands
            print("Printing all commands:")
            print("T1 - generate data for excel and db")
            print("T2 -  update data for db")
        elif mainInput == "T1":
            print("Enter from which year the data should be generated")
            yearOfStart = int(input())

            print("Enter until which year the data should be generated")
            yearOfEnd = int(input())

            print("Enter number of courses you want to generate")
            numberOfRoutes = int(input())

            numOfControllers = int(20 + floor(numberOfRoutes / 1000))
            numOfDrivers = int(20 + floor(numberOfRoutes / 1000))

            generateControllers(yearOfStart, yearOfEnd, numOfControllers)
            generateDrivers(yearOfStart, yearOfEnd, numOfDrivers)

            workBook.remove_sheet(workBook["Sheet"])
            workBook.save("WorkBook.xlsx")
            workBook.save("WorkBook1.xlsx")

            # DB
            numOfBuses = 10 + floor(numberOfRoutes / 10000)
            getBuses(numOfBuses)  # 10 buses required, +1 for each 10000 courses

            getCoursesAndRides(yearOfStart, yearOfEnd, numberOfRoutes, numOfControllers, numOfBuses, controllersExcel, driversExcel, numOfDrivers)

            print("Generating finished")
        elif mainInput == "T2":
            print("Enter from which year the data should be generated")
            yearOfStart = int(input())

            print("Enter until which year the data should be generated")
            yearOfEnd = int(input())

            print("Enter number of courses you want to generate")
            numberOfRoutes = int(input())

            numOfControllers = int(2 + floor(numberOfRoutes / 1000))
            numOfDrivers = int(2 + floor(numberOfRoutes / 1000))

            # Excel
            fireDrivers(0.15, yearOfStart, yearOfEnd)
            fireControlers(0.15, yearOfStart, yearOfEnd)

            generateControllers(yearOfStart, yearOfEnd, numOfControllers)
            generateDrivers(yearOfStart, yearOfEnd, numOfDrivers)

            # DB
            getCoursesAndRides(yearOfStart, yearOfEnd, numberOfRoutes, numOfControllers, 0, controllersExcel, driversExcel,
                               numOfDrivers, True)


            workBook.save("WorkBook2.xlsx")
            if os.path.exists("WorkBook.xlsx"):
                os.remove("WorkBook.xlsx")

            print("Generating finished")
        else:
            print("Write either T1 or T2")


if os.path.exists("WorkBook2.xlsx"):
    os.remove("WorkBook2.xlsx")
if os.path.exists("WorkBook1.xlsx"):
    os.remove("WorkBook1.xlsx")


workBook = openpyxl.Workbook()
ws1 = workBook.create_sheet("Controllers")
ws2 = workBook.create_sheet("Drivers")
wsPayroles = []
mainLoop()
